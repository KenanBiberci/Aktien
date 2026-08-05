"""
excel.py — Excel-Output (Schritt 6).

Erzeugt ein Workbook mit vier Blättern (Screener, Annahmen, Methodik, Top-Ideen),
bedingter Formatierung auf dem Signal, sichtbarem Disclaimer und optionalen
Detailblättern je Ticker (--details).

Font Arial; Zahlenformate: Preise #,##0.00, Prozente 0.0%, Multiplikatoren 0.0x.
"""
from __future__ import annotations

import math
from datetime import datetime
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- Formate -----------------------------------------------------------------
FONT = "Arial"
FMT_PRICE = "#,##0.00"
FMT_PCT = "0.0%"
FMT_MULT = "0.0x"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # Eingaben deutlich markiert
TITLE_FONT = Font(name=FONT, bold=True, size=14)
NORMAL_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")

DISCLAIMER = (
    "Dieses Tool wendet Standard-Bewertungsmethoden mechanisch an. Die Signale sind "
    "regelbasiert und nur so gut wie die (kostenlosen) Daten und die gewählten Annahmen. "
    "Uniforme/Sektor-Median-Multiplikatoren sind Näherungen. KEINE ANLAGEBERATUNG. "
    "Vor jeder Entscheidung eigene Prüfung und Risikostreuung. „Enormes Potenzial\" = hohes "
    "Risiko; bei Hebelprodukten (z. B. Knock-outs) droht Totalverlust."
)

# Screener-Spalten: (Überschrift, DataFrame-Spalte, Zahlenformat)
SCREENER_COLS: list[tuple[str, str, str | None]] = [
    ("Ticker", "yahoo", None),
    ("Name", "security", None),
    ("Sektor", "sector", None),
    ("Kurs", "price", FMT_PRICE),
    ("KGV ttm", "kgv_ttm", FMT_MULT),
    ("KGV fwd", "kgv_fwd", FMT_MULT),
    ("Div-Rendite", "div_yield", FMT_PCT),
    ("Payout", "payout", FMT_PCT),
    ("ROE", "roe", FMT_PCT),
    ("M1 Gordon", "m1", FMT_PRICE),
    ("M2 DDM", "m2", FMT_PRICE),
    ("M3 Justified PE", "m3", FMT_PRICE),
    ("M4 Comp-PE", "m4", FMT_PRICE),
    ("M5 P/S", "m5", FMT_PRICE),
    ("M6 P/B", "m6", FMT_PRICE),
    ("M7 EV/EBITDA", "m7", FMT_PRICE),
    ("M8 DCF", "m8", FMT_PRICE),
    ("#Methoden", "n_methods", None),
    ("Blended Fair Value", "blended_fair_value", FMT_PRICE),
    ("Blended Upside", "blended_upside", FMT_PCT),
    ("Konsens-Upside", "consensus_upside", FMT_PCT),
    ("Ø-Upside", "avg_upside", FMT_PCT),
    ("BUY-SIGNAL", "signal", None),
    ("Analysten-Konsens", "rec_key", None),
    ("PVGO %", "pvgo_pct", FMT_PCT),
    ("Divergenz", "divergence", FMT_PCT),
]

SIGNAL_COL_HEADER = "BUY-SIGNAL"


def _cell_value(val: Any) -> Any:
    """NaN/inf/None -> None (leer); sonst der Wert."""
    if val is None:
        return None
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return None
    if isinstance(val, (np.floating,)):
        f = float(val)
        return None if (math.isnan(f) or math.isinf(f)) else f
    if isinstance(val, (np.integer,)):
        return int(val)
    return val


def build_workbook(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    medians: dict[str, dict[str, float]],
    run_date: datetime,
    details: list[str] | None = None,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # Default-Blatt weg

    _sheet_screener(wb.create_sheet("Screener"), df)
    _sheet_screener(wb.create_sheet("Screener (Ø-Upside)"),
                    df.sort_values("avg_upside", ascending=False, na_position="last"))
    _sheet_assumptions(wb.create_sheet("Annahmen"), cfg, medians)
    _sheet_methodik(wb.create_sheet("Methodik"), cfg, run_date)
    _sheet_top_ideas(wb.create_sheet("Top-Ideen"), df)

    if details:
        for ticker in details:
            sub = df[df["yahoo"] == ticker]
            if not sub.empty:
                _sheet_detail(wb.create_sheet(f"Detail {ticker}"[:31]),
                              sub.iloc[0].to_dict(), cfg)
    return wb


# =============================================================================
# a) Screener
# =============================================================================
def _sheet_screener(ws: Worksheet, df: pd.DataFrame) -> None:
    headers = [c[0] for c in SCREENER_COLS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    signal_col_idx = headers.index(SIGNAL_COL_HEADER) + 1

    for r, (_, rowdata) in enumerate(df.iterrows(), start=2):
        for col_idx, (_, field, fmt) in enumerate(SCREENER_COLS, start=1):
            val = _cell_value(rowdata.get(field))
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.border = BORDER
            if fmt and isinstance(val, (int, float)):
                cell.number_format = fmt
        _color_signal_cell(ws.cell(row=r, column=signal_col_idx))

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    _add_signal_conditional_formatting(ws, signal_col_idx, last_row)
    _autosize(ws, headers)


def _add_signal_conditional_formatting(ws: Worksheet, col_idx: int, last_row: int) -> None:
    """Echte bedingte Formatierung auf der Signal-Spalte (grün/gelb/rot)."""
    if last_row < 2:
        return
    letter = get_column_letter(col_idx)
    rng = f"{letter}2:{letter}{last_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"STRONG BUY"'], fill=GREEN))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"BUY"'], fill=GREEN))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"HOLD"'], fill=YELLOW))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"REDUCE"'], fill=RED))


def _color_signal_cell(cell: Any) -> None:
    val = str(cell.value or "")
    if val in ("STRONG BUY", "BUY"):
        cell.fill = GREEN
    elif val == "HOLD":
        cell.fill = YELLOW
    elif val == "REDUCE":
        cell.fill = RED
    cell.font = Font(name=FONT, size=10, bold=val in ("STRONG BUY", "REDUCE"))


# =============================================================================
# b) Annahmen
# =============================================================================
def _sheet_assumptions(ws: Worksheet, cfg: dict[str, Any],
                       medians: dict[str, dict[str, float]]) -> None:
    ws["A1"] = "Annahmen (Eingaben — gelb hinterlegt, ohne Codeänderung anpassbar)"
    ws["A1"].font = TITLE_FONT

    rows: list[tuple[str, Any, str | None]] = [
        ("Risk-free rate (rf)", cfg["risk_free_rate"], FMT_PCT),
        ("Equity risk premium (ERP)", cfg["equity_risk_premium"], FMT_PCT),
        ("Terminal growth (g)", cfg["terminal_growth"], FMT_PCT),
        ("Stage-1 Jahre (N)", cfg["stage1_years"], None),
        ("DDM Payout-Gate", cfg["ddm_payout_gate"], FMT_PCT),
        ("Steuersatz", cfg["default_tax"], FMT_PCT),
        ("Signal: STRONG BUY ab", cfg["signal_thresholds"]["strong_buy"], FMT_PCT),
        ("Signal: BUY ab", cfg["signal_thresholds"]["buy"], FMT_PCT),
        ("Signal: HOLD-Floor (darunter REDUCE)", cfg["signal_thresholds"]["hold_floor"], FMT_PCT),
        ("Fallback PE", cfg["fallback_multiples"]["pe"], FMT_MULT),
        ("Fallback P/S", cfg["fallback_multiples"]["ps"], FMT_MULT),
        ("Fallback P/B", cfg["fallback_multiples"]["pb"], FMT_MULT),
        ("Fallback EV/EBITDA", cfg["fallback_multiples"]["ev_ebitda"], FMT_MULT),
    ]
    r = 3
    ws.cell(row=r, column=1, value="Globaler Hebel").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=2, value="Wert").font = HEADER_FONT
    ws.cell(row=r, column=2).fill = HEADER_FILL
    r += 1
    for label, value, fmt in rows:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        cell = ws.cell(row=r, column=2, value=_cell_value(value))
        cell.fill = INPUT_FILL
        cell.font = NORMAL_FONT
        if fmt:
            cell.number_format = fmt
        r += 1

    # Sektor-Median-Multiplikatoren-Tabelle
    r += 1
    ws.cell(row=r, column=1, value="Berechnete Sektor-Median-Multiplikatoren "
            "(faire Multiplikatoren der Comparable-Methoden)").font = TITLE_FONT
    r += 1
    med_headers = ["GICS-Sektor", "PE (fwd)", "P/S", "P/B", "EV/EBITDA"]
    for c_idx, h in enumerate(med_headers, start=1):
        cell = ws.cell(row=r, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    r += 1
    for sector in sorted(medians.keys()):
        m = medians[sector]
        ws.cell(row=r, column=1, value=sector).font = NORMAL_FONT
        for c_idx, key in enumerate(("pe", "ps", "pb", "ev_ebitda"), start=2):
            cell = ws.cell(row=r, column=c_idx, value=_cell_value(m.get(key)))
            cell.number_format = FMT_MULT
            cell.font = NORMAL_FONT
        r += 1

    ws.column_dimensions["A"].width = 45
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 16


# =============================================================================
# c) Methodik
# =============================================================================
def _sheet_methodik(ws: Worksheet, cfg: dict[str, Any], run_date: datetime) -> None:
    ws["A1"] = "Methodik & Disclaimer"
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 120

    lines = [
        f"Datum des Laufs: {run_date.strftime('%Y-%m-%d %H:%M')}",
        "Datenquellen: yfinance (primär), optional Financial Modeling Prep (FMP_API_KEY). "
        "S&P-500-Liste: Wikipedia mit Fallback auf datasets/s-and-p-500-companies.",
        "",
        "Die 9 Methoden (jeweils fairer Preis je Aktie):",
        "  M1 Gordon-Growth: D1/(r-g), nur wenn dps>0, r>g, payout>=Gate.",
        "  M2 2-Stufen-DDM: N Jahre g1, danach g; nur wenn dps>0, r>g1, r>g, payout>=Gate.",
        "  M3 Fundamentales KGV: justified_pe = payout/(r-g); x EPS_fwd; nur wenn payout>=Gate.",
        "  M4 Comparable-KGV: Sektor-Median-PE x EPS_fwd.",
        "  M5 P/S: Sektor-Median-P/S x Umsatz je Aktie.",
        "  M6 P/B: Sektor-Median-P/B x Buchwert je Aktie.",
        "  M7 EV/EBITDA: Sektor-Median x EBITDA -> Equity je Aktie = (EV - Net Debt)/Shares.",
        "  M8 DCF/FCFF: FCFF=(OpCF - Capex), wachsende Perpetuität; guard wacc>g.",
        "  M9 Asset-based: Buchwert je Aktie (nur Info-Spalte).",
        "",
        "Eigenkapitalkosten r (CAPM): r = rf + beta x ERP (beta fehlt -> 1).",
        "WACC: (E/(E+D)) x r + (D/(E+D)) x r_d x (1-Steuer), r_d = rf + 1,5%.",
        "g1 (Stufe 1): clip(ROE x (1-payout), g, 20%).",
        "",
        "Blend & Signal:",
        "  Blended Fair Value = Median gültiger Methoden (M1-M3 nur wenn payout>=Gate; M4-M8 immer).",
        f"  STRONG BUY: Ø-Upside >= {cfg['signal_thresholds']['strong_buy']:.0%} und #Methoden>=3.",
        f"  BUY: Ø-Upside >= {cfg['signal_thresholds']['buy']:.0%}.",
        f"  HOLD: Ø-Upside >= {cfg['signal_thresholds']['hold_floor']:.0%}.",
        "  REDUCE: darunter. N/A – Datenlücke: #Methoden<2 und kein Analystenziel.",
        "  Ø-Upside = Mittel aus Blended-Upside und Konsens-Upside (soweit vorhanden).",
        "",
        "Ergänzend: PVGO = Kurs - EPS_ttm/r (Barwert der Wachstumschancen); "
        "Value Creation = Ja wenn ROE > r; Divergenz = max/min der Methoden - 1.",
        "",
        "── DISCLAIMER ──",
        DISCLAIMER,
    ]
    for i, line in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("── DISCLAIMER") or "KEINE ANLAGEBERATUNG" in line:
            cell.font = Font(name=FONT, size=11, bold=True, color="C00000")


# =============================================================================
# d) Top-Ideen
# =============================================================================
def _sheet_top_ideas(ws: Worksheet, df: pd.DataFrame) -> None:
    top = df[df["signal"].isin(["STRONG BUY", "BUY"])].copy()
    top = top.sort_values("avg_upside", ascending=False, na_position="last")

    headers = ["Ticker", "Name", "Sektor", "Kurs", "Ø-Upside", "Signal",
               "Analysten-Konsens", "Kurz-Notiz"]
    fields = ["yahoo", "security", "sector", "price", "avg_upside", "signal", "rec_key", None]
    fmts = [None, None, None, FMT_PRICE, FMT_PCT, None, None, None]

    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, (_, rowdata) in enumerate(top.iterrows(), start=2):
        for c_idx, (field, fmt) in enumerate(zip(fields, fmts), start=1):
            if field is None:
                continue
            val = _cell_value(rowdata.get(field))
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = NORMAL_FONT
            if fmt and isinstance(val, (int, float)):
                cell.number_format = fmt
        _color_signal_cell(ws.cell(row=r, column=6))

    ws.freeze_panes = "A2"
    _autosize(ws, headers)


# =============================================================================
# Detailblatt (optional, --details)
# =============================================================================
def _sheet_detail(ws: Worksheet, row: dict[str, Any], cfg: dict[str, Any]) -> None:
    ws["A1"] = f"Detail: {row.get('security')} ({row.get('yahoo')})"
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    methods = [
        ("M1 Gordon-Growth", "m1"), ("M2 2-Stufen-DDM", "m2"),
        ("M3 Fundamentales KGV", "m3"), ("M4 Comparable-KGV", "m4"),
        ("M5 P/S", "m5"), ("M6 P/B", "m6"), ("M7 EV/EBITDA", "m7"),
        ("M8 DCF/FCFF", "m8"), ("M9 Asset-based (Info)", "m9"),
    ]
    r = 3
    ws.cell(row=r, column=1, value="Methode").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=2, value="Fairer Preis").font = HEADER_FONT
    ws.cell(row=r, column=2).fill = HEADER_FILL
    r += 1
    for label, key in methods:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        cell = ws.cell(row=r, column=2, value=_cell_value(row.get(key)))
        cell.number_format = FMT_PRICE
        cell.font = NORMAL_FONT
        r += 1

    r += 1
    extras = [
        ("Kurs", row.get("price"), FMT_PRICE),
        ("Blended Fair Value", row.get("blended_fair_value"), FMT_PRICE),
        ("Ø-Upside", row.get("avg_upside"), FMT_PCT),
        ("Signal", row.get("signal"), None),
        ("PVGO %", row.get("pvgo_pct"), FMT_PCT),
        ("Value Creation (ROE>r)", row.get("value_creation"), None),
        ("r (CAPM)", row.get("r"), FMT_PCT),
        ("WACC", row.get("wacc"), FMT_PCT),
    ]
    for label, value, fmt in extras:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        cell = ws.cell(row=r, column=2, value=_cell_value(value))
        if fmt:
            cell.number_format = fmt
        cell.font = NORMAL_FONT
        r += 1

    # r/g-Sensitivitätsraster: P0 = D1/(r-g)
    r += 1
    ws.cell(row=r, column=1, value="Sensitivität P0 = D1/(r-g)").font = TITLE_FONT
    r += 1
    g = float(cfg["terminal_growth"])
    dps = float(row.get("dps") or 0.0)
    d1 = dps * (1 + g)
    r_base = float(row.get("r") or cfg["risk_free_rate"])
    r_values = [r_base - 0.01, r_base, r_base + 0.01, r_base + 0.02]
    g_values = [g - 0.01, g, g + 0.01]

    ws.cell(row=r, column=1, value="r \\ g").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    for c_idx, gv in enumerate(g_values, start=2):
        cell = ws.cell(row=r, column=c_idx, value=gv)
        cell.number_format = FMT_PCT
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    r += 1
    for rv in r_values:
        cell = ws.cell(row=r, column=1, value=rv)
        cell.number_format = FMT_PCT
        cell.font = NORMAL_FONT
        for c_idx, gv in enumerate(g_values, start=2):
            price = d1 / (rv - gv) if (rv > gv and dps > 0) else None
            cc = ws.cell(row=r, column=c_idx, value=price)
            cc.number_format = FMT_PRICE
            cc.font = NORMAL_FONT
        r += 1


# =============================================================================
# Hilfen
# =============================================================================
def _autosize(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        width = max(10, min(28, len(str(header)) + 2))
        ws.column_dimensions[letter].width = width
